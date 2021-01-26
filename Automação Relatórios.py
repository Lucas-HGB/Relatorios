#!/usr/bin/env python
# -*- coding: utf-8 -*- 
from os import mkdir
from shutil import copyfile
from pandas import read_excel
from collections import Counter
from datetime import datetime, timedelta
from pyzabbix import ZabbixAPI
from requests import get
from io import BytesIO
from PIL.Image import open as pil_open
from time import mktime
from Filter import Convert
from openpyxl import load_workbook, Workbook


global cookies, headers, time_start, time_end, workbook

## Seta tempo para ser usado como padrao para extrair valores de items
time_end = int(mktime(datetime.now().timetuple()))
time_start = time_end - 60 * 60 * 24 * 31


try:
	workbook = load_workbook("Item_Values.xlsx")
except IOError:
	workbook = Workbook()
except Exception as excp:
	print (excp)

## Cookies e headers necessários para autenticar site do zabbix e baixar imagens dos gráficos (Removidos do Github por motivos de privacidade.
cookies = {
    'PHPSESSID': '',
    'zbx_sessionid': '',
}

headers = {
    'Connection': '',
    'Cache-Control': '',
    'Upgrade-Insecure-Requests': '',
    'Origin': '',
    'Content-Type': '',
    'User-Agent': '',
    'Accept': '',
    'Referer': '',
    'Accept-Language': '',
    'If-None-Match': '',
    'Content-Length': '',
    'X-Requested-With': '',
}

global situacao_servidores
situacao_servidores = {}
global zabbix

## Loga na API do zabbix (usado para extrair id dos gráficos e também items/valores)
zabbix = ZabbixAPI("http://guardiao.workdb.com.br")
zabbix.login("lucas.hoeltgebaum", "workdb#2020")

class Item():

    def __init__(self, id, servidor):
        self.id = id
        self.servidor = servidor
        self.has_data = True
        self.name = zabbix.item.get(itemids = self.id)[0]["name"]
		
    def extract_history(self):
        ## Extrai histórico do item
        history = zabbix.history.get(itemids=[self.id], time_from=time_start, time_till=time_end, output='extend', limit='10000000')
        if not len(history):
            history = zabbix.history.get(itemids=[self.id], time_from=time_start, time_till=time_end, output='extend', limit='10000000', history=0)
        ## Separa apenas os valores do item em uma variável diferente
        values = [float(f["value"]) for f in history]
        ## Se valor não for nulo
        if values != [] and values[0] != 0:
            self.last = history[-1]["value"]
            self.min = min(values)
            self.max = max(values)
            self.med = int((sum(values) / len(values)))
            ## Passa valores para o filtro, que transforma Bytes em sua devida unidade
            self.last, self.max, self.med, self.min = Convert(self.name, self.last, self.max, self.med, self.min)
        ## Se valor for nulo
        else:
            self.has_data = False


    def save(self):
        ## Se valor não for nulo, item é salvo na planilha
        if self.has_data:
            sheet = Sheet(self.servidor)
            sheet.basic_setup(self.id, self.name, self.last, self.max, self.med, self.min)
        else:
            pass

class Graph():

    def __init__(self, id, servidor, graph_name):
        self.id = id
        self.servidor = servidor
        self.name = graph_name


    ## Remove caracteres inválidos que causam erros ao salvar arquivo em sistemas Windows
    def remove_invalid_char(self, name):
        words_blacklist = [
        "{$SID}", 
        "-", 
        "/", 
        "{}".format("/"), 
        "(SEG)", 
        "º"
        ]
        filtered_word = ""
        altered = False
        for word in name.split():
            if "#" in word:
                if not altered:
                    filtered_word = name.replace("#", "")
                elif altered:
                    filtered_word = filtered_word.replace("#", "")
                altered = True
            elif word in words_blacklist:
                if not altered:
                    filtered_word = name.replace(word, "")
                elif altered:
                    filtered_word = filtered_word.replace(word, "")
                altered = True
            elif "/" in word:
                if not altered:
                    filtered_word = name.replace("/", "")
                elif altered:
                    filtered_word = filtered_word.replace("/", "")
                altered = True
            elif ":" in word:
                if not altered:
                    filtered_word = name.replace(":", "")
                elif altered:
                    filtered_word = filtered_word.replace(":", "")
                altered = True
            else:
                filtered_word = name
            filtered_word = filtered_word
        while filtered_word[-1] == " ":
            filtered_word = filtered_word[0:-1]
        return filtered_word

    def get_img(self):
        ## Baixa bytes da imagem do site do zabbix, utilizando cookies e headers definidos no início como autentição com o guardião
        if "swap" not in self.name.lower() and "disk" not in self.name.lower() and "grupo" not in self.name.lower():
            response = get('http://guardiao.workdb.com.br/chart2.php?graphid={}&from=now-1M%2FM&to=now-1M%2FM&profileIdx=web.graphs.filter&profileIdx2={}=um5etv25&screenid='.format(self.id, self.id), headers=headers, cookies=cookies, verify=False)
        else:
            response = get('http://guardiao.workdb.com.br/chart2.php?graphid={}&from=now-1M%2FM&to=now-1M%2FM&profileIdx=web.graphs.filter&profileIdx2={}&width=1274&height=280&_=um5ge3fh&screenid='.format(self.id, self.id), headers=headers, cookies=cookies, verify=False)
        ## Converte bytes extraidos do site em uma imagem
        img = pil_open(BytesIO(response.content))
        ## Filtra nome do gráfico e então salva a imagem
        name = self.remove_invalid_char(self.name)
        img.save(r"Relatórios\{}\{}.png".format(self.servidor, name))
        

class Servidor():
    def __init__(self, servidor, id):
        self.servidor = servidor
        self.id = id
        ## Extrai gráficos e items de respectivo servidor
        self.graphs = (zabbix.graph.get(hostids = self.id))
        self.items = zabbix.item.get(hostids = self.id)
        try:
            mkdir("Relatórios")
        except FileExistsError:
            pass
        try:
            ## Cria diretório do servidor em questão
            mkdir(r"Relatórios\{}".format(self.servidor))
        except FileExistsError:
            pass

    def move_model(self):
        ## Move modelo do arquivo Word para o folder do servidor
        try:
            try:
                copyfile(r"Modelos\{}\_Model.docx".format(self.servidor), r"Relatórios\{}\_Model.docx".format(self.servidor))
            except FileNotFoundError:
                print("Model {} not in correct folder!".format(self.servidor))
            except FileExistsError:
                print("Model {} already in folder with relatório".format(self.servidor))
        except Exception as excp:
            print(excp)

    def get_graphs(self):
        count = 0
        ## Para cada gráfico extraído da API do zabbix, extrai a imagem do site do guardião
        for graph in self.graphs:
            count += 1
            graph_obj = Graph(id = graph["graphid"], servidor = self.servidor, graph_name = graph["name"])
            graph_obj.get_img()
            print(f"Saved {count}/{len(self.graphs)} images.")

    def get_values(self):
        count = 0
        ## Para cada item extraído da API do zabbix, extrai os valores
        for item_id in self.items:
            count += 1
            item = Item(item_id["itemid"], self.servidor)
            item.extract_history() ; item.save()
            print(f"Extracted {count}/{len(self.items)} item values.")

            

class Sheet():
	def __init__(self, servidor):
        ## Cria planilha para servidor especificado
		try:
			self.worksheet = workbook[servidor]
		except:
			self.worksheet = workbook.create_sheet(servidor)
			self.worksheet = workbook[servidor]

	def basic_setup(self, id, name, last, max, med, min):
        ## Salva valores passados na planilha de devido servidor
		rows = [("{}".format( name), "Last: {}".format(last), "Max: {}".format(max), "Med: {}".format(med), "Min: {}".format(min))]
		for f in rows:
			self.worksheet.append(f)


def START(type = None, name = None, id = None):
    if type != None:
        if type.lower() == "workdb":
            servidores_excel = read_excel("Servidores.xlsx", sheet_name = "WorkDB")
        elif type.lower() == "optidata":
            servidores_excel = read_excel("Servidores.xlsx", sheet_name = "Optidata")
        ## Para cada servidor especificado na planilha, roda todo o processo
        for server, id, count in zip(servidores_excel["SERVER"], servidores_excel["ID"], range(len(servidores_excel["SERVER"]))):
            print("\n\nProcessed {}/{} Servers.".format(count, len(servidores_excel["SERVER"])))
            print(f"Starting data collection from server {server} with ID {id}\n")
            Servidor_obj = Servidor(server, id)
            Servidor_obj.move_model()
            Servidor_obj.get_graphs()
            Servidor_obj.get_values()
            
            
    elif type == None:
        print(f"\n\nStarting data collection from server {name} with ID {id}")
        Servidor_obj = Servidor(name, id)
        Servidor_obj.move_model()
        Servidor_obj.get_graphs()
        Servidor_obj.get_values()

exit = False

while not exit:
    print("1 - Gerar relatórios Optidata")
    print("2 - Gerar relatórios WorkDB")
    print("3 - Gerar relatório de servidor específico")
    print("4 - Sair")
    opcao = input()
    try:
        opcao = int(opcao)
    except ValueError:
        print("Insira um valor entre 1-3")
        opcao = input()
    if opcao == 1:
        START(type = "optidata")
    elif opcao == 2:
        START(type = "workdb")
    elif opcao == 3:
        print("Nome:")
        nome = input()
        print("ID:")
        try:
            id = int(input())
            START(type = None, name = nome.upper(), id = id)
        except ValueError:
            pass
        
    elif opcao == 4:
        break
    else:
        print("Por favor insira um valor válido!")

## Salva planilha com valores extraídos dos items
workbook.save(r"Item_Values.xlsx")
